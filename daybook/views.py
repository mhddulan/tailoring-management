from datetime import date, datetime
from decimal import Decimal
from django.utils import timezone
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from branches.models import Branch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from .forms import (
    DayBookForm,
    OpeningBalanceForm,
)
from .models import DayBook, OpeningBalance


# ============================================================
# HELPERS
# ============================================================

ZERO = Decimal("0.00")


def get_user_branches(request):

    if request.user.role == "Admin":
        return Branch.objects.all().order_by("name")

    return Branch.objects.filter(
        id=request.user.branch_id
    )


def get_selected_branch(request):

    if request.user.role == "Admin":

        branch_id = request.GET.get("branch")

        if branch_id:
            return Branch.objects.filter(
                id=branch_id
            ).first()

        return None

    return request.user.branch


def get_base_entries(request):

    if request.user.role == "Admin":

        return DayBook.objects.select_related(
            "branch"
        ).all()

    return DayBook.objects.select_related(
        "branch"
    ).filter(
        branch=request.user.branch
    )


def get_period_entries(request):

    entries = get_base_entries(request)

    branch = request.GET.get("branch")

    if request.user.role == "Admin" and branch:

        entries = entries.filter(
            branch_id=branch
        )

    from_date = request.GET.get("from_date")

    to_date = request.GET.get("to_date")

    if from_date:
        entries = entries.filter(
            date__gte=from_date
        )

    if to_date:
        entries = entries.filter(
            date__lte=to_date
        )

    transaction_type = request.GET.get(
        "transaction_type"
    )

    if transaction_type in [
        "Income",
        "Expense",
    ]:
        entries = entries.filter(
            transaction_type=transaction_type
        )

    payment_mode = request.GET.get(
        "payment_mode"
    )

    if payment_mode:
        entries = entries.filter(
            payment_mode=payment_mode
        )

    category = request.GET.get(
        "category"
    )

    if category:
        entries = entries.filter(
            category=category
        )

    return entries.order_by(
        "-date",
        "-id"
    )


def get_previous_entries(request, from_date):

    if not from_date:
        return DayBook.objects.none()

    entries = get_base_entries(request)

    branch = request.GET.get("branch")

    if request.user.role == "Admin" and branch:

        entries = entries.filter(
            branch_id=branch
        )

    return entries.filter(
        date__lt=from_date
    )


def calculate_summary(entries):

    income = (
        entries
        .filter(transaction_type="Income")
        .aggregate(
            total=Sum("amount")
        )["total"]
        or ZERO
    )

    expense = (
        entries
        .filter(transaction_type="Expense")
        .aggregate(
            total=Sum("amount")
        )["total"]
        or ZERO
    )

    cash_income = (
        entries
        .filter(
            transaction_type="Income",
            payment_mode="Cash"
        )
        .aggregate(
            total=Sum("amount")
        )["total"]
        or ZERO
    )

    cash_expense = (
        entries
        .filter(
            transaction_type="Expense",
            payment_mode="Cash"
        )
        .aggregate(
            total=Sum("amount")
        )["total"]
        or ZERO
    )

    # Bank + Online + POS are treated as bank balance
    bank_income = (
        entries
        .filter(
            transaction_type="Income",
            payment_mode__in=[
                "Bank",
                "Online",
                "POS",
            ]
        )
        .aggregate(
            total=Sum("amount")
        )["total"]
        or ZERO
    )

    bank_expense = (
        entries
        .filter(
            transaction_type="Expense",
            payment_mode__in=[
                "Bank",
                "Online",
                "POS",
            ]
        )
        .aggregate(
            total=Sum("amount")
        )["total"]
        or ZERO
    )

    return {
        "income": income,
        "expense": expense,
        "profit": income - expense,

        "cash_income": cash_income,
        "cash_expense": cash_expense,

        "bank_income": bank_income,
        "bank_expense": bank_expense,
    }


def get_opening_balances(request, from_date):

    """
    Returns opening Cash and Bank balances.

    If a date is selected:
        Opening = configured opening balance
                 + all transactions before that date.

    If no date is selected:
        Opening = configured opening balance.
    """

    branches = get_user_branches(request)

    selected_branch_id = request.GET.get("branch")

    # --------------------------------------------------------
    # Admin + specific branch
    # --------------------------------------------------------

    if request.user.role == "Admin" and selected_branch_id:

        branch = Branch.objects.filter(
            id=selected_branch_id
        ).first()

        if not branch:
            return ZERO, ZERO

        opening = OpeningBalance.objects.filter(
            branch=branch
        ).first()

        opening_cash = (
            opening.opening_cash
            if opening
            else ZERO
        )

        opening_bank = (
            opening.opening_bank
            if opening
            else ZERO
        )

        previous_entries = DayBook.objects.filter(
            branch=branch
        )

        if from_date:

            previous_entries = previous_entries.filter(
                date__lt=from_date
            )

    # --------------------------------------------------------
    # Branch Manager
    # --------------------------------------------------------

    elif request.user.role != "Admin":

        branch = request.user.branch

        if not branch:
            return ZERO, ZERO

        opening = OpeningBalance.objects.filter(
            branch=branch
        ).first()

        opening_cash = (
            opening.opening_cash
            if opening
            else ZERO
        )

        opening_bank = (
            opening.opening_bank
            if opening
            else ZERO
        )

        previous_entries = DayBook.objects.filter(
            branch=branch
        )

        if from_date:

            previous_entries = previous_entries.filter(
                date__lt=from_date
            )

    # --------------------------------------------------------
    # Admin + ALL branches
    # --------------------------------------------------------

    else:

        opening_records = OpeningBalance.objects.filter(
            branch__in=branches
        )

        opening_cash = (
            opening_records.aggregate(
                total=Sum("opening_cash")
            )["total"]
            or ZERO
        )

        opening_bank = (
            opening_records.aggregate(
                total=Sum("opening_bank")
            )["total"]
            or ZERO
        )

        previous_entries = DayBook.objects.filter(
            branch__in=branches
        )

        if from_date:

            previous_entries = previous_entries.filter(
                date__lt=from_date
            )

    # --------------------------------------------------------
    # Previous CASH
    # --------------------------------------------------------

    previous_cash_income = (
        previous_entries
        .filter(
            transaction_type="Income",
            payment_mode="Cash"
        )
        .aggregate(
            total=Sum("amount")
        )["total"]
        or ZERO
    )

    previous_cash_expense = (
        previous_entries
        .filter(
            transaction_type="Expense",
            payment_mode="Cash"
        )
        .aggregate(
            total=Sum("amount")
        )["total"]
        or ZERO
    )

    # --------------------------------------------------------
    # Previous BANK
    # --------------------------------------------------------

    previous_bank_income = (
        previous_entries
        .filter(
            transaction_type="Income",
            payment_mode__in=[
                "Bank",
                "Online",
                "POS",
            ]
        )
        .aggregate(
            total=Sum("amount")
        )["total"]
        or ZERO
    )

    previous_bank_expense = (
        previous_entries
        .filter(
            transaction_type="Expense",
            payment_mode__in=[
                "Bank",
                "Online",
                "POS",
            ]
        )
        .aggregate(
            total=Sum("amount")
        )["total"]
        or ZERO
    )

    opening_cash = (
        opening_cash
        + previous_cash_income
        - previous_cash_expense
    )

    opening_bank = (
        opening_bank
        + previous_bank_income
        - previous_bank_expense
    )

    return opening_cash, opening_bank


# ============================================================
# DAY BOOK LIST
# ============================================================
@login_required
def daybook_list(request):

    today = timezone.localdate()

    # =====================================================
    # USER BRANCH
    # =====================================================

    if request.user.role == "Branch":

        branch = request.user.branch

        entries = DayBook.objects.filter(
            branch=branch
        )

        branches = Branch.objects.filter(
            id=branch.id
        )

    else:

        entries = DayBook.objects.all()

        branches = Branch.objects.all()


    # =====================================================
    # DATE FILTER
    # =====================================================

    from_date = request.GET.get("from_date", "").strip()
    to_date = request.GET.get("to_date", "").strip()


    # Default = Today

    if not from_date:

        from_date = today

    else:

        try:
            from_date = timezone.datetime.strptime(
                from_date,
                "%Y-%m-%d"
            ).date()

        except ValueError:

            from_date = today


    if not to_date:

        to_date = from_date

    else:

        try:
            to_date = timezone.datetime.strptime(
                to_date,
                "%Y-%m-%d"
            ).date()

        except ValueError:

            to_date = from_date


    # =====================================================
    # BRANCH FILTER - ADMIN
    # =====================================================

    branch_id = request.GET.get(
        "branch",
        ""
    ).strip()


    if request.user.role == "Admin" and branch_id:

        entries = entries.filter(
            branch_id=branch_id
        )


    # =====================================================
    # TRANSACTION TYPE
    # =====================================================

    transaction_type = request.GET.get(
        "transaction_type",
        ""
    ).strip()


    if transaction_type:

        entries = entries.filter(
            transaction_type=transaction_type
        )


    # =====================================================
    # PAYMENT MODE
    # =====================================================

    payment_mode = request.GET.get(
        "payment_mode",
        ""
    ).strip()


    if payment_mode:

        entries = entries.filter(
            payment_mode=payment_mode
        )


    # =====================================================
    # CATEGORY
    # =====================================================

    category = request.GET.get(
        "category",
        ""
    ).strip()


    if category:

        entries = entries.filter(
            category=category
        )


    # =====================================================
    # DATE FILTER
    # =====================================================

    filtered_entries = entries.filter(
        date__range=[
            from_date,
            to_date
        ]
    ).order_by(
        "-date",
        "-id"
    )


    # =====================================================
    # SELECTED PERIOD INCOME
    # =====================================================

    income = filtered_entries.filter(
        transaction_type="Income"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    # =====================================================
    # SELECTED PERIOD EXPENSE
    # =====================================================

    expense = filtered_entries.filter(
        transaction_type="Expense"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    # =====================================================
    # SELECTED PERIOD PROFIT
    # =====================================================

    profit = income - expense


    # =====================================================
    # MONTHLY TOTALS
    # =====================================================

    month_start = today.replace(
        day=1
    )


    monthly_entries = entries.filter(
        date__range=[
            month_start,
            today
        ]
    )


    monthly_income = monthly_entries.filter(
        transaction_type="Income"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    monthly_expense = monthly_entries.filter(
        transaction_type="Expense"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    monthly_profit = (
        monthly_income -
        monthly_expense
    )


    # =====================================================
    # CASH - SELECTED PERIOD
    # =====================================================

    cash_income = filtered_entries.filter(
        transaction_type="Income",
        payment_mode="Cash"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    cash_expense = filtered_entries.filter(
        transaction_type="Expense",
        payment_mode="Cash"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    # =====================================================
    # BANK - SELECTED PERIOD
    # =====================================================

    bank_income = filtered_entries.filter(
        transaction_type="Income",
        payment_mode="Bank"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    bank_expense = filtered_entries.filter(
        transaction_type="Expense",
        payment_mode="Bank"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    # =====================================================
    # OPENING CASH
    # =====================================================

    opening_cash = entries.filter(
        date__lt=from_date,
        payment_mode="Cash"
    ).aggregate(
        income=Sum(
            "amount",
            filter=None
        )
    )


    cash_before_income = entries.filter(
        date__lt=from_date,
        payment_mode="Cash",
        transaction_type="Income"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    cash_before_expense = entries.filter(
        date__lt=from_date,
        payment_mode="Cash",
        transaction_type="Expense"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    opening_cash = (
        cash_before_income -
        cash_before_expense
    )


    # =====================================================
    # OPENING BANK
    # =====================================================

    bank_before_income = entries.filter(
        date__lt=from_date,
        payment_mode="Bank",
        transaction_type="Income"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    bank_before_expense = entries.filter(
        date__lt=from_date,
        payment_mode="Bank",
        transaction_type="Expense"
    ).aggregate(
        total=Sum("amount")
    )["total"] or 0


    opening_bank = (
        bank_before_income -
        bank_before_expense
    )


    # =====================================================
    # CLOSING CASH
    # =====================================================

    closing_cash = (
        opening_cash +
        cash_income -
        cash_expense
    )


    # =====================================================
    # CLOSING BANK
    # =====================================================

    closing_bank = (
        opening_bank +
        bank_income -
        bank_expense
    )


    # =====================================================
    # TOTAL BALANCE
    # =====================================================

    total_closing_balance = (
        closing_cash +
        closing_bank
    )


    # =====================================================
    # CONTEXT
    # =====================================================

    context = {

        "entries": filtered_entries,

        "branches": branches,

        "today": today,

        "from_date": from_date,

        "to_date": to_date,

        "income": income,

        "expense": expense,

        "profit": profit,

        "monthly_income": monthly_income,

        "monthly_expense": monthly_expense,

        "monthly_profit": monthly_profit,

        "cash_income": cash_income,

        "cash_expense": cash_expense,

        "bank_income": bank_income,

        "bank_expense": bank_expense,

        "opening_cash": opening_cash,

        "opening_bank": opening_bank,

        "closing_cash": closing_cash,

        "closing_bank": closing_bank,

        "total_closing_balance": total_closing_balance,

    }


    return render(
        request,
        "daybook/daybook_list.html",
        context
    )

# ============================================================
# ADD DAY BOOK
# ============================================================
@login_required
def daybook_create(request):

    if request.method == "POST":

        form = DayBookForm(
            request.POST,
            user=request.user
        )

        if form.is_valid():

            entry = form.save(
                commit=False
            )

            # Branch user → automatically use own branch
            if request.user.role != "Admin":

                entry.branch = request.user.branch

            # Admin must have a branch
            if request.user.role == "Admin" and not entry.branch_id:

                form.add_error(
                    "branch",
                    "Please select a branch."
                )

            else:

                entry.save()

                messages.success(
                    request,
                    "Day Book entry added successfully."
                )

                return redirect(
                    "daybook_list"
                )

        # IMPORTANT:
        # Print validation errors in terminal
        print("====================================")
        print("DAY BOOK FORM ERRORS")
        print(form.errors)
        print("====================================")

    else:

        form = DayBookForm(
            user=request.user
        )

    return render(
        request,
        "daybook/daybook_form.html",
        {
            "form": form
        }
    )


# ============================================================
# EDIT DAY BOOK
# ============================================================

@login_required
def daybook_edit(request, id):

    if request.user.role == "Admin":

        entry = get_object_or_404(
            DayBook,
            id=id
        )

    else:

        entry = get_object_or_404(
            DayBook,
            id=id,
            branch=request.user.branch
        )

    if request.method == "POST":

        form = DayBookForm(
            request.POST,
            instance=entry,
            user=request.user
        )

        if form.is_valid():

            updated_entry = form.save(
                commit=False
            )

            if request.user.role != "Admin":

                updated_entry.branch = (
                    request.user.branch
                )

            updated_entry.save()

            messages.success(
                request,
                "Day Book entry updated successfully."
            )

            return redirect(
                "daybook_list"
            )

    else:

        form = DayBookForm(
            instance=entry,
            user=request.user
        )

    return render(
        request,
        "daybook/daybook_form.html",
        {
            "form": form,
            "entry": entry
        }
    )


# ============================================================
# DELETE DAY BOOK
# ============================================================

@login_required
def daybook_delete(request, id):

    if request.user.role == "Admin":

        entry = get_object_or_404(
            DayBook,
            id=id
        )

    else:

        entry = get_object_or_404(
            DayBook,
            id=id,
            branch=request.user.branch
        )

    if request.method == "POST":

        entry.delete()

        messages.success(
            request,
            "Day Book entry deleted successfully."
        )

        return redirect(
            "daybook_list"
        )

    return render(
        request,
        "daybook/daybook_delete.html",
        {
            "entry": entry
        }
    )


# ============================================================
# EXCEL EXPORT
# ============================================================

@login_required
def daybook_excel(request):

    entries = get_period_entries(request)

    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    opening_cash, opening_bank = get_opening_balances(
        request,
        from_date
    )

    summary = calculate_summary(entries)

    closing_cash = (
        opening_cash
        + summary["cash_income"]
        - summary["cash_expense"]
    )

    closing_bank = (
        opening_bank
        + summary["bank_income"]
        - summary["bank_expense"]
    )

    total_closing = (
        closing_cash
        + closing_bank
    )

    net_profit = (
        summary["income"]
        - summary["expense"]
    )

    # ========================================================
    # WORKBOOK
    # ========================================================

    workbook = Workbook()

    ws = workbook.active
    ws.title = "Day Book"

    ws.sheet_view.showGridLines = False

    # ========================================================
    # STYLES
    # ========================================================

    title_font = Font(
        bold=True,
        size=18
    )

    subtitle_font = Font(
        bold=True,
        size=13
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    bold_font = Font(
        bold=True
    )

    thin = Side(
        style="thin",
        color="B7B7B7"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    right = Alignment(
        horizontal="right",
        vertical="center"
    )

    # ========================================================
    # BRANCH NAME
    # ========================================================

    branch_id = request.GET.get("branch")

    if request.user.role == "Admin" and branch_id:

        branch = Branch.objects.filter(
            id=branch_id
        ).first()

        branch_name = (
            branch.name
            if branch
            else "Unknown"
        )

    elif request.user.role == "Admin":

        branch_name = "All Branches"

    else:

        branch_name = (
            request.user.branch.name
            if request.user.branch
            else "Unknown"
        )

    # ========================================================
    # TITLE
    # ========================================================

    ws.merge_cells("A1:G1")

    ws["A1"] = "STITCHING PRO"

    ws["A1"].font = title_font

    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")

    ws["A2"] = "DAY BOOK ACCOUNTING REPORT"

    ws["A2"].font = subtitle_font

    ws["A2"].alignment = center

    ws.merge_cells("A3:G3")

    ws["A3"] = f"Branch: {branch_name}"

    ws["A3"].alignment = center

    # ========================================================
    # REPORT INFORMATION
    # ========================================================

    ws["A5"] = "From Date"
    ws["B5"] = from_date or "All"

    ws["D5"] = "To Date"
    ws["E5"] = to_date or "All"

    ws["A6"] = "Generated"
    ws["B6"] = datetime.now().strftime(
        "%d-%m-%Y %H:%M"
    )

    for cell in [
        "A5",
        "D5",
        "A6",
    ]:

        ws[cell].font = bold_font

    # ========================================================
    # ACCOUNT SUMMARY
    # ========================================================

    ws.merge_cells("A8:G8")

    ws["A8"] = "ACCOUNT SUMMARY"

    ws["A8"].font = subtitle_font

    # Headers

    summary_headers = [
        "Account",
        "Opening",
        "Income",
        "Expense",
        "Closing",
    ]

    for col, value in enumerate(
        summary_headers,
        start=1
    ):

        cell = ws.cell(
            row=9,
            column=col,
            value=value
        )

        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Cash

    summary_rows = [

        [
            "Cash in Hand",
            opening_cash,
            summary["cash_income"],
            summary["cash_expense"],
            closing_cash,
        ],

        [
            "Cash in Bank",
            opening_bank,
            summary["bank_income"],
            summary["bank_expense"],
            closing_bank,
        ],

        [
            "Total Balance",
            opening_cash + opening_bank,
            summary["income"],
            summary["expense"],
            total_closing,
        ],
    ]

    row = 10

    for data in summary_rows:

        for col, value in enumerate(
            data,
            start=1
        ):

            cell = ws.cell(
                row=row,
                column=col,
                value=value
            )

            cell.border = border

            if col == 1:

                cell.font = bold_font

            else:

                cell.number_format = (
                    'SAR #,##0.00'
                )

                cell.alignment = right

        row += 1

    # ========================================================
    # PROFIT SUMMARY
    # ========================================================

    ws["A14"] = "PROFIT SUMMARY"

    ws["A14"].font = subtitle_font

    profit_rows = [

        ["Total Income", summary["income"]],

        ["Total Expense", summary["expense"]],

        ["Net Profit", net_profit],

    ]

    row = 15

    for label, amount in profit_rows:

        ws.cell(
            row=row,
            column=1,
            value=label
        ).font = bold_font

        ws.cell(
            row=row,
            column=2,
            value=amount
        )

        ws.cell(
            row=row,
            column=2
        ).number_format = (
            'SAR #,##0.00'
        )

        row += 1

    # ========================================================
    # TRANSACTION TABLE
    # ========================================================

    table_start = 20

    headers = [
        "Date",
        "Branch",
        "Type",
        "Category",
        "Payment Mode",
        "Description",
        "Amount",
    ]

    for col, header in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=table_start,
            column=col,
            value=header
        )

        cell.font = header_font
        cell.alignment = center
        cell.border = border

    current_row = table_start + 1

    for entry in entries:

        values = [

            entry.date,

            str(entry.branch),

            entry.transaction_type,

            entry.category,

            entry.payment_mode,

            entry.description or "-",

            entry.amount,

        ]

        for col, value in enumerate(
            values,
            start=1
        ):

            cell = ws.cell(
                row=current_row,
                column=col,
                value=value
            )

            cell.border = border

            if col == 7:

                cell.number_format = (
                    'SAR #,##0.00'
                )

                cell.alignment = right

        current_row += 1

    # ========================================================
    # TRANSACTION TOTALS
    # ========================================================

    total_income = sum(
        (
            entry.amount
            for entry in entries
            if entry.transaction_type == "Income"
        ),
        ZERO
    )

    total_expense = sum(
        (
            entry.amount
            for entry in entries
            if entry.transaction_type == "Expense"
        ),
        ZERO
    )

    ws.cell(
        row=current_row + 1,
        column=6,
        value="TOTAL INCOME"
    ).font = bold_font

    ws.cell(
        row=current_row + 1,
        column=7,
        value=total_income
    )

    ws.cell(
        row=current_row + 1,
        column=7
    ).number_format = 'SAR #,##0.00'

    ws.cell(
        row=current_row + 2,
        column=6,
        value="TOTAL EXPENSE"
    ).font = bold_font

    ws.cell(
        row=current_row + 2,
        column=7,
        value=total_expense
    )

    ws.cell(
        row=current_row + 2,
        column=7
    ).number_format = 'SAR #,##0.00'

    ws.cell(
        row=current_row + 3,
        column=6,
        value="NET PROFIT"
    ).font = bold_font

    ws.cell(
        row=current_row + 3,
        column=7,
        value=total_income - total_expense
    )

    ws.cell(
        row=current_row + 3,
        column=7
    ).number_format = 'SAR #,##0.00'

    # ========================================================
    # COLUMN WIDTHS
    # ========================================================

    widths = {
        "A": 16,
        "B": 22,
        "C": 16,
        "D": 25,
        "E": 20,
        "F": 45,
        "G": 20,
    }

    for column, width in widths.items():

        ws.column_dimensions[
            column
        ].width = width

    # ========================================================
    # FREEZE
    # ========================================================

    ws.freeze_panes = (
        f"A{table_start + 1}"
    )

    # ========================================================
    # PRINT SETTINGS
    # ========================================================

    ws.page_setup.orientation = "landscape"

    ws.page_setup.fitToWidth = 1

    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.print_title_rows = (
        f"{table_start}:{table_start}"
    )

    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # ========================================================
    # RESPONSE
    # ========================================================

    response = HttpResponse(
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        'attachment; filename="StitchingPro_DayBook.xlsx"'
    )

    workbook.save(response)

    return response
# ============================================================
# PDF EXPORT
# ============================================================

@login_required
def daybook_pdf(request):

    entries = get_period_entries(request)

    from_date = request.GET.get("from_date")

    to_date = request.GET.get("to_date")

    opening_cash, opening_bank = (
        get_opening_balances(
            request,
            from_date
        )
    )

    summary = calculate_summary(entries)

    closing_cash = (
        opening_cash
        + summary["cash_income"]
        - summary["cash_expense"]
    )

    closing_bank = (
        opening_bank
        + summary["bank_income"]
        - summary["bank_expense"]
    )

    total_closing = (
        closing_cash
        + closing_bank
    )

    # --------------------------------------------------------
    # RESPONSE
    # --------------------------------------------------------

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; filename="DayBook_Report.pdf"'
    )

    pdf = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "Title",
        parent=styles["Title"],
        alignment=TA_CENTER,
        fontSize=18,
        leading=22,
        spaceAfter=4,
    )

    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        fontSize=11,
        spaceAfter=10,
    )

    right_style = ParagraphStyle(
        "Right",
        parent=styles["Normal"],
        alignment=TA_RIGHT,
    )

    story = []

    # --------------------------------------------------------
    # HEADER
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "STITCHING PRO",
            title_style
        )
    )

    story.append(
        Paragraph(
            "DAY BOOK ACCOUNTING REPORT",
            subtitle_style
        )
    )

    branch_id = request.GET.get("branch")

    if request.user.role == "Admin" and branch_id:

        branch = Branch.objects.filter(
            id=branch_id
        ).first()

        branch_name = (
            branch.name
            if branch
            else "Unknown"
        )

    elif request.user.role == "Admin":

        branch_name = "All Branches"

    else:

        branch_name = (
            request.user.branch.name
            if request.user.branch
            else "Unknown"
        )

    report_info = [
        [
            Paragraph(
                f"<b>Branch:</b> {branch_name}",
                styles["Normal"]
            ),
            Paragraph(
                f"<b>From:</b> {from_date or 'All'}",
                styles["Normal"]
            ),
            Paragraph(
                f"<b>To:</b> {to_date or 'All'}",
                styles["Normal"]
            ),
            Paragraph(
                f"<b>Generated:</b> "
                f"{datetime.now().strftime('%d-%m-%Y %H:%M')}",
                styles["Normal"]
            ),
        ]
    ]

    info_table = Table(
        report_info,
        colWidths=[
            65 * mm,
            50 * mm,
            50 * mm,
            65 * mm,
        ]
    )

    info_table.setStyle(
        TableStyle([
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                6
            ),
        ])
    )

    story.append(info_table)

    story.append(
        Spacer(1, 5)
    )

    # --------------------------------------------------------
    # ACCOUNT SUMMARY
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "<b>ACCOUNT SUMMARY</b>",
            styles["Heading3"]
        )
    )

    summary_data = [

        [
            "Cash in Hand",
            "Opening",
            "Income",
            "Expense",
            "Closing",
        ],

        [
            "",
            f"SAR {opening_cash:,.2f}",
            f"SAR {summary['cash_income']:,.2f}",
            f"SAR {summary['cash_expense']:,.2f}",
            f"SAR {closing_cash:,.2f}",
        ],

        [
            "Cash in Bank",
            f"SAR {opening_bank:,.2f}",
            f"SAR {summary['bank_income']:,.2f}",
            f"SAR {summary['bank_expense']:,.2f}",
            f"SAR {closing_bank:,.2f}",
        ],

        [
            "TOTAL",
            "",
            "",
            "",
            f"SAR {total_closing:,.2f}",
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[
            45 * mm,
            38 * mm,
            38 * mm,
            38 * mm,
            45 * mm,
        ]
    )

    summary_table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#343a40")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "ALIGN",
                (1, 0),
                (-1, -1),
                "RIGHT"
            ),

            (
                "FONTNAME",
                (0, 3),
                (-1, 3),
                "Helvetica-Bold"
            ),

            (
                "BACKGROUND",
                (0, 3),
                (-1, 3),
                colors.HexColor("#e9ecef")
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                6
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                6
            ),
        ])
    )

    story.append(
        summary_table
    )

    story.append(
        Spacer(1, 10)
    )

    # --------------------------------------------------------
    # TRANSACTIONS
    # --------------------------------------------------------

    story.append(
        Paragraph(
            "<b>TRANSACTION DETAILS</b>",
            styles["Heading3"]
        )
    )

    data = [[
        "Date",
        "Branch",
        "Type",
        "Category",
        "Payment",
        "Description",
        "Amount",
    ]]

    for entry in entries:

        data.append([

            entry.date.strftime(
                "%d-%m-%Y"
            ),

            str(entry.branch),

            entry.transaction_type,

            entry.category,

            entry.payment_mode,

            entry.description or "-",

            f"SAR {entry.amount:,.2f}",

        ])

    table = Table(
        data,
        repeatRows=1,
        colWidths=[
            25 * mm,
            35 * mm,
            25 * mm,
            40 * mm,
            28 * mm,
            70 * mm,
            32 * mm,
        ]
    )

    table.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#343a40")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP"
            ),

            (
                "ALIGN",
                (-1, 1),
                (-1, -1),
                "RIGHT"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                5
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                5
            ),
        ])
    )

    story.append(
        table
    )

    story.append(
        Spacer(1, 10)
    )

    story.append(
        Paragraph(
            f"<b>Total Closing Balance: "
            f"SAR {total_closing:,.2f}</b>",
            right_style
        )
    )

    # --------------------------------------------------------
    # FOOTER
    # --------------------------------------------------------

    def add_page_number(canvas, doc):

        canvas.saveState()

        canvas.setFont(
            "Helvetica",
            8
        )

        canvas.drawString(
            10 * mm,
            6 * mm,
            "Stitching Pro - Day Book"
        )

        canvas.drawRightString(
            landscape(A4)[0] - 10 * mm,
            6 * mm,
            f"Page {doc.page}"
        )

        canvas.restoreState()

    pdf.build(
        story,
        onFirstPage=add_page_number,
        onLaterPages=add_page_number,
    )

    return response


# ============================================================
# FILTER FUNCTION
# ============================================================

def get_filtered_entries(request):

    return get_period_entries(request)
@login_required
def opening_balance(request):

    # Only Admin can manage opening balances
    if request.user.role != "Admin":
        messages.error(
            request,
            "Only Admin can manage opening balances."
        )
        return redirect("daybook_list")

    branch_id = request.GET.get("branch")

    instance = None

    if branch_id:
        instance = OpeningBalance.objects.filter(
            branch_id=branch_id
        ).first()

    if request.method == "POST":

        branch_id = request.POST.get("branch")

        instance = OpeningBalance.objects.filter(
            branch_id=branch_id
        ).first()

        form = OpeningBalanceForm(
            request.POST,
            instance=instance,
            user=request.user
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Opening Cash and Bank balance saved successfully."
            )

            return redirect(
                "opening_balance"
            )

    else:

        form = OpeningBalanceForm(
            instance=instance,
            user=request.user
        )

    return render(
        request,
        "daybook/opening_balance.html",
        {
            "form": form,
            "branches": Branch.objects.all().order_by("name"),
        }
    )